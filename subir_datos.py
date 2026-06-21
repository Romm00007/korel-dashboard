import openpyxl, json, requests, base64
from datetime import datetime

# ── CAMBIA ESTOS 2 VALORES ──────────────────────────────
EXCEL_FILE   = 'FORMATO_DASHBOARD.xlsx'
GITHUB_TOKEN = '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'
GITHUB_REPO  = 'Romm00007/korel-dashboard'
# ────────────────────────────────────────────────────────

def excel_a_json():
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]
        result[name] = [
            {k: (v.isoformat() if hasattr(v, 'isoformat') else v)
             for k, v in zip(headers, row)}
            for row in rows[1:] if any(c is not None for c in row)
        ]
    wb.close()
    result['_actualizado'] = datetime.now().isoformat()
    return result

def subir_a_github(data):
    url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/data.json'
    headers = {'Authorization': f'token {GITHUB_TOKEN}'}
    r = requests.get(url, headers=headers)
    sha = r.json().get('sha') if r.status_code == 200 else None
    contenido = base64.b64encode(json.dumps(data, ensure_ascii=False, indent=2).encode()).decode()
    payload = {'message': f'Update {datetime.now().strftime("%H:%M")}', 'content': contenido}
    if sha:
        payload['sha'] = sha
    r = requests.put(url, headers=headers, json=payload)
    if r.status_code in (200, 201):
        print(f'✅ Subido — {datetime.now().strftime("%H:%M:%S")}')
    else:
        print(f'❌ Error: {r.status_code}')

data = excel_a_json()
subir_a_github(data)
